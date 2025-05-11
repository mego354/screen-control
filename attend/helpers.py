import os, time
from django.conf import settings
from django.utils import timezone
from datetime import timedelta

from django_user_agents.utils import get_user_agent
from django.core.mail import EmailMessage

from openpyxl.styles import Alignment
from openpyxl import Workbook

from django.template.loader import render_to_string

from pathlib import Path
 
def get_user_device(request):
        user_agent =  get_user_agent(request)
        device_info = {
            "is_mobile": user_agent.is_mobile,
            "browser_family": user_agent.browser.family,
            "os_family": user_agent.os.family,
            "device_model": user_agent.device.model,
            "device_family": (user_agent.device.family if user_agent.device.family != "Other" else None),
        }

        return device_info

def close_event(event, close_minuts):
    time.sleep(60 * close_minuts)
    event.is_active = False
    event.is_done = True
    event.save()

def delete_oldest_files(folder_path, keep_count):
    folder = Path(folder_path)    
    files = [f for f in folder.iterdir() if f.is_file()]
    files.sort(key=lambda x: x.stat().st_mtime)  # Sort by modification time (oldest first)
    if len(files) > keep_count:
        files_to_delete = files[:-keep_count] if keep_count > 0 else files
        for file in files_to_delete:
            os.remove(file)

def generate_exc(event):
    model_name = event.__class__.__name__
    destination = os.path.join(settings.MEDIA_ROOT, "excels", model_name)
    if not os.path.exists(destination):
        os.makedirs(destination)
    filename = event.excel_filename()
    file_path = os.path.join(destination, filename)

    if os.path.isfile(file_path):
        return file_path
    else:
        if model_name == 'Lecture':
            audiance = f"Group: {event.group.number}"
        elif model_name == 'Section':
            audiance = f"Group: {event.group.number} - Class: {event.classs.number}"
        wb = Workbook()
        ws1 = wb.active
        ws1.sheet_properties.tabColor = "00FF00"

        ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
        ws1.merge_cells(start_row=2, start_column=1, end_row=2, end_column=3)
        ws1.merge_cells(start_row=3, start_column=1, end_row=3, end_column=5)
        ws1.merge_cells(start_row=4, start_column=1, end_row=4, end_column=5)

        ws1['A1'] = f"{event.DepCourse.course.name}"
        ws1['A2'] = f"Date: {event.date.strftime('%d/%m/%y')}"
        ws1['A3'] = audiance
        ws1["D1"] = "Type:"
        ws1["E1"] = model_name
        ws1["E1"].alignment = Alignment(horizontal='center')
        ws1["D2"] = "audiance:"
        ws1["E2"] = event.attendees.count()
        ws1["E2"].alignment = Alignment(horizontal='center')
        ws1["A5"] = "Number"
        ws1["C5"] = "Name"
        ws1["E5"] = "ID"

        attendees = list(
        event.attendees.all()
        .values_list("id", "first_name", "last_name")
        .order_by("id")
        )
        row_start = 6
        col_start = 1
        for index, (id, first_name, last_name) in enumerate(attendees):
            ws1.cell(row=row_start + index, column=col_start).value = index + 1
            ws1.merge_cells(start_row=row_start + index, start_column=col_start + 1, end_row=row_start + index, end_column=col_start + 3)
            name_cell = ws1.cell(row=row_start + index, column=col_start + 1)
            name_cell.value = f"{first_name} {last_name}"

            name_cell.alignment = Alignment(horizontal='center')
            ws1.cell(row=row_start + index, column=col_start + 4).value = id


        wb.save(file_path)

        today_date = timezone.localtime().date()
        
        events_count = event.__class__.objects.filter(date__range=(today_date - timedelta(days=1), today_date)).count()
        delete_oldest_files(destination, events_count + 10)

        return file_path



def send_event_email(event, file_path):
    model_name = event.__class__.__name__
    date = timezone.localtime(event.date)
    formatted_date = date.strftime("%d-%m-%y %H:%M")
    if model_name == 'Lecture':
        audiance = f"Group: {event.group.number}"
        email_subject = f"Attendance of lecture {event.DepCourse.course.name} G{event.group.number}"
    elif model_name == 'Section':
        audiance = f"Group: {event.group.number} - Class: {event.classs.number}"
        email_subject = f"Attendance of lab {event.DepCourse.course.name} G{event.group.number} C{event.classs.number}"

    context = {
        'doctor_name': str(event.doctor),
        'audiance': audiance,
        'date': formatted_date,
    }

    body = render_to_string('events/event_email.html', context)

    to_email = event.doctor.user.email
    if not to_email or to_email == "":
        email_addresses = ['megomego354@gmail.com']
        body = "NOT DONE " + body
    else:
        email_addresses = ['megomego354@gmail.com', to_email]

    email = EmailMessage(
        subject=email_subject,
        body=body,
        from_email=settings.DEFAULT_FROM_EMAIL,
        to=email_addresses
    )

    # Attach the Excel file to the email
    email.attach_file(file_path)
    email.send()




